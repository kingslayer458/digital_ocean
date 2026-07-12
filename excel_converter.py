import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def json_to_excel(*json_files):
    """
    Convert one or more JSON files to Excel files.
    The Excel file will have the same name as the JSON file.
    """

    for json_file in json_files:
        excel_file = os.path.splitext(json_file)[0] + ".xlsx"

        df = pd.read_json(json_file)

        # Remove timezone from datetime columns
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)

        df.to_excel(excel_file, index=False)

        # Auto-adjust column widths
        wb = load_workbook(excel_file)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_file)

        print(f" Created: {excel_file}")

if __name__ == "__main__":
    json_to_excel("vm_data_digital_ocean.json")